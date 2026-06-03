"""
Cloud sync for the unified tender portal.

Fetches both live in-progress tenders and submitted tenders from the Tender
site, then writes a single root data.json consumed by Cloudflare Pages.
"""

import hashlib
import json
import logging
import os
import time
from datetime import datetime
from io import BytesIO

import openpyxl
import requests
from bs4 import BeautifulSoup
from msal import PublicClientApplication


logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

ALRAWAF_URL = "https://www.alrawaf-ms.com"
ALRAWAF_USER = os.getenv("ALRAWAF_USERNAME")
ALRAWAF_PASS = os.getenv("ALRAWAF_PASSWORD")
ALRAWAF_CLIENT_ID = os.getenv("ALRAWAF_CLIENT_ID", "8e8af720-b54b-4f52-8858-95109b4a2c5d")
ALRAWAF_TENANT = os.getenv("ALRAWAF_TENANT", "alrawaf.com.sa")
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "data.json")

# Supabase (تُضبط كأسرار في GitHub Actions)
SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")


def request_with_retry(session, method, url, attempts=4, **kwargs):
    last_error = None
    retry_statuses = {408, 429, 500, 502, 503, 504}
    for attempt in range(1, attempts + 1):
        try:
            response = session.request(method, url, **kwargs)
            if response.status_code not in retry_statuses:
                return response
            last_error = requests.HTTPError(f"HTTP {response.status_code}", response=response)
            log.warning("Transient HTTP %s on %s attempt %s/%s", response.status_code, url, attempt, attempts)
        except requests.RequestException as exc:
            last_error = exc
            log.warning("Transient request error on %s attempt %s/%s: %s", url, attempt, attempts, exc)
        if attempt < attempts:
            time.sleep(min(20, 2 ** attempt))
    if last_error:
        raise last_error
    raise RuntimeError(f"Request failed: {url}")


class AlrawafClient:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        })
        self._token = None

    def get_token(self):
        if self._token:
            return self._token

        app = PublicClientApplication(
            ALRAWAF_CLIENT_ID,
            authority=f"https://login.microsoftonline.com/{ALRAWAF_TENANT}",
        )
        scopes_to_try = [
            [f"api://{ALRAWAF_CLIENT_ID}/user_impersonation"],
            ["https://graph.microsoft.com/User.Read"],
        ]

        for scopes in scopes_to_try:
            result = app.acquire_token_by_username_password(
                username=ALRAWAF_USER,
                password=ALRAWAF_PASS,
                scopes=scopes,
            )
            if "access_token" in result:
                self._token = result["access_token"]
                log.info("Microsoft token acquired for Tender site")
                return self._token
            log.warning("Token attempt failed: %s", result.get("error_description", "")[:240])

        return None

    def auth_session(self):
        token = self.get_token()
        if token:
            self.session.headers.update({"Authorization": f"Bearer {token}"})

    def fetch_excel(self, url, row_map):
        try:
            response = request_with_retry(self.session, "GET", url, timeout=30)
            if response.status_code != 200 or len(response.content) < 100:
                log.warning("Excel download failed from %s: %s", url, response.status_code)
                return []

            workbook = openpyxl.load_workbook(BytesIO(response.content), read_only=True, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            log.info("Excel headers from %s: %s", url.rstrip("/").split("/")[-1], headers)

            def to_str(value):
                if value is None:
                    return ""
                return str(value).replace(" 00:00:00", "").strip()

            tenders = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                title_idx = row_map.get("اسم المناقصة", 3)
                title = to_str(row[title_idx]) if len(row) > title_idx else ""
                if not title:
                    continue

                tender = {"tender_id": hashlib.md5(title.encode("utf-8")).hexdigest()[:12]}
                for field, idx in row_map.items():
                    tender[field] = to_str(row[idx]) if len(row) > idx else ""
                tender["fetched_at"] = datetime.now().isoformat()
                tenders.append(tender)

            log.info("Fetched %s rows from %s", len(tenders), url)
            return tenders
        except Exception as exc:
            log.error("Excel fetch error from %s: %s", url, exc)
            return []

    def fetch_in_progress(self):
        self.auth_session()
        row_map = {
            "اسم المناقصة": 3,
            "تاريخ التقديم": 5,
            "تاريخ الضمان الابتدائي": 4,
            "المالك": 6,
            "نوع الأعمال": 7,
            "القطاع": 8,
            "الحالة": 9,
        }
        tenders = self.fetch_excel(
            f"{ALRAWAF_URL}/en/tendering/export_to_excel_in_progress_tender/",
            row_map,
        )
        for tender in tenders:
            tender["الحالة"] = "جارية"
        return tenders

    def fetch_submitted(self):
        self.auth_session()
        row_map = {
            "اسم المناقصة": 3,
            "تاريخ التقديم": 5,
            "تاريخ الضمان الابتدائي": 4,
            "المالك": 6,
            "نوع الأعمال": 7,
            "القطاع": 8,
        }
        export_urls = [
            f"{ALRAWAF_URL}/en/tendering/export_to_excel_submitted_tender/",
            f"{ALRAWAF_URL}/en/tendering/export_submitted/",
            f"{ALRAWAF_URL}/en/tendering/tender_report_submitted/export/",
        ]

        for url in export_urls:
            submitted = self.fetch_excel(url, row_map)
            if submitted:
                for tender in submitted:
                    tender["الحالة"] = "تم التقديم"
                return submitted

        return self.fetch_submitted_from_page()

    def fetch_submitted_from_page(self):
        try:
            response = request_with_retry(
                self.session,
                "GET",
                f"{ALRAWAF_URL}/en/tendering/tender_report_submitted/",
                timeout=30,
            )
            if response.status_code != 200:
                log.warning("Submitted page fetch failed: %s", response.status_code)
                return []

            soup = BeautifulSoup(response.text, "html.parser")
            export_link = soup.find("a", href=lambda href: href and "export" in href.lower() and "excel" in href.lower())
            if export_link:
                export_url = ALRAWAF_URL + export_link["href"]
                row_map = {
                    "اسم المناقصة": 3,
                    "تاريخ التقديم": 5,
                    "تاريخ الضمان الابتدائي": 4,
                    "المالك": 6,
                    "نوع الأعمال": 7,
                    "القطاع": 8,
                }
                submitted = self.fetch_excel(export_url, row_map)
                for tender in submitted:
                    tender["الحالة"] = "تم التقديم"
                return submitted

            tenders = []
            rows = soup.select("tr.o_data_row, table tbody tr, .o_list_view tbody tr")
            for row in rows:
                cells = row.find_all("td")
                if len(cells) < 2:
                    continue
                title = cells[0].get_text(strip=True)
                if not title:
                    continue
                tenders.append({
                    "tender_id": hashlib.md5(title.encode("utf-8")).hexdigest()[:12],
                    "اسم المناقصة": title,
                    "تاريخ التقديم": cells[1].get_text(strip=True) if len(cells) > 1 else "",
                    "المالك": cells[2].get_text(strip=True) if len(cells) > 2 else "",
                    "نوع الأعمال": cells[3].get_text(strip=True) if len(cells) > 3 else "",
                    "القطاع": cells[4].get_text(strip=True) if len(cells) > 4 else "",
                    "تاريخ الضمان الابتدائي": cells[5].get_text(strip=True) if len(cells) > 5 else "",
                    "الحالة": "تم التقديم",
                    "fetched_at": datetime.now().isoformat(),
                })
            return tenders
        except Exception as exc:
            log.error("Submitted page fetch error: %s", exc)
            return []


def _safe_date(value):
    """يحوّل نصاً إلى تاريخ ISO صالح (YYYY-MM-DD) أو None."""
    if not value:
        return None
    text = str(value).strip()[:10]
    try:
        datetime.fromisoformat(text)
        return text
    except ValueError:
        return None


def push_to_supabase(all_tenders):
    """يرفع/يحدّث المناقصات في جدول tenders عبر Supabase REST (upsert).

    لا يحذف المناقصات المفقودة من المصدر كي نحافظ على المهام والتعيينات
    والاعتمادات المرتبطة بها داخل لوحة العمليات.
    """
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        log.info("Supabase env not set — skipping database push (data.json only).")
        return

    rows = []
    for tender in all_tenders:
        tender_id = tender.get("tender_id") or tender.get("id")
        if not tender_id:
            continue
        rows.append({
            "id": tender_id,
            "title": tender.get("اسم المناقصة") or "",
            "client": tender.get("المالك") or "",
            "sector": tender.get("القطاع") or "",
            "work_type": tender.get("نوع الأعمال") or "",
            "submit_date": _safe_date(tender.get("تاريخ التقديم")),
            "guarantee_date": _safe_date(tender.get("تاريخ الضمان الابتدائي")),
            "external_status": tender.get("الحالة") or "",
            "fetched_at": tender.get("fetched_at"),
        })

    if not rows:
        log.warning("No rows to push to Supabase.")
        return

    endpoint = f"{SUPABASE_URL}/rest/v1/tenders?on_conflict=id"
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    try:
        response = requests.post(endpoint, headers=headers, data=json.dumps(rows), timeout=30)
        if response.status_code in (200, 201, 204):
            log.info("Pushed %s tenders to Supabase (upsert).", len(rows))
        else:
            log.error("Supabase push failed: %s %s", response.status_code, response.text[:300])
    except requests.RequestException as exc:
        log.error("Supabase push error: %s", exc)


def main():
    missing = [name for name in ["ALRAWAF_USERNAME", "ALRAWAF_PASSWORD"] if not os.getenv(name)]
    if missing:
        raise SystemExit(f"Missing required secrets: {', '.join(missing)}")

    client = AlrawafClient()
    in_progress = client.fetch_in_progress()
    submitted = client.fetch_submitted()

    if not in_progress and not submitted:
        raise SystemExit("No tender data was fetched; keeping existing data.json unchanged.")

    data = {
        "last_updated": datetime.now().isoformat(),
        "source": "alrawaf-tender-live",
        "tenders": in_progress,
        "submitted": submitted,
    }
    with open(OUTPUT_FILE, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
        handle.write("\n")

    log.info("Wrote %s with %s in-progress and %s submitted tenders", OUTPUT_FILE, len(in_progress), len(submitted))

    # حافظ على تزامن قاعدة بيانات Supabase مع المصدر
    push_to_supabase(in_progress + submitted)


if __name__ == "__main__":
    main()
