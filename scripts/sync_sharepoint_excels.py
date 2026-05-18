"""
Sync SharePoint Excel workbooks into JSON files consumed by the portal.

This script is designed for GitHub Actions. It uses Microsoft Graph app-only
authentication, downloads the two shared Excel workbooks, and writes:

- portfolio/data.json
- executive-report/data.json
"""

import base64
import json
import logging
import os
import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import openpyxl
import requests


logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

GRAPH_ROOT = "https://graph.microsoft.com/v1.0"
VAT_RATE = 0.15

DEFAULT_EXECUTIVE_URL = (
    "https://etech321.sharepoint.com/:x:/s/e-tech/"
    "IQCz-YQi0yYuTo2nSWtkJ_nFAZhUxo9BGN-5lxd8Zewcs6Q?e=szVgi8"
)
DEFAULT_PORTFOLIO_URL = (
    "https://etech321.sharepoint.com/:x:/s/e-tech/"
    "IQBUS72YnIkdSbJgXENbslnUAUGJqGmiN8KSVH3EH8KTvvQ?e=b5McqU"
)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def graph_token() -> str:
    missing = [name for name in ("SP_TENANT_ID", "SP_CLIENT_ID", "SP_CLIENT_SECRET") if not os.getenv(name)]
    if missing:
        raise RuntimeError(f"Missing required GitHub secrets: {', '.join(missing)}")
    tenant_id = os.environ["SP_TENANT_ID"]
    client_id = os.environ["SP_CLIENT_ID"]
    client_secret = os.environ["SP_CLIENT_SECRET"]
    response = requests.post(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data={
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["access_token"]


def share_id(shared_url: str) -> str:
    encoded = base64.urlsafe_b64encode(shared_url.encode("utf-8")).decode("ascii").rstrip("=")
    return f"u!{encoded}"


def graph_get(url: str, token: str) -> requests.Response:
    response = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    response.raise_for_status()
    return response


def download_shared_workbook(shared_url: str, token: str) -> bytes:
    sid = share_id(shared_url)
    meta = graph_get(f"{GRAPH_ROOT}/shares/{sid}/driveItem", token).json()
    log.info("Downloading workbook: %s", meta.get("name", "unknown.xlsx"))
    return graph_get(f"{GRAPH_ROOT}/shares/{sid}/driveItem/content", token).content


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).replace("\r\n", "\n").strip()


def number_value(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    text = re.sub(r"[^\d.\-]", "", text)
    try:
        return float(text)
    except ValueError:
        return 0.0


def norm_key(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"\s+", " ", text)
    return text


def find_col(headers: list[Any], *needles: str, fallback: int | None = None) -> int | None:
    normalized = [norm_key(h) for h in headers]
    for needle in needles:
        n = needle.lower()
        for idx, header in enumerate(normalized):
            if n in header:
                return idx
    return fallback


def find_amount_col(headers: list[Any], include_vat: bool) -> int | None:
    normalized = [norm_key(h) for h in headers]
    for idx, header in enumerate(normalized):
        if not any(word in header for word in ("قيمة", "مبلغ", "اجمالي", "الإجمالي", "الاجمالي")):
            continue
        has_vat = "شامل" in header and "غير شامل" not in header
        excludes_vat = "غير شامل" in header or "بدون ضريبة" in header
        if include_vat and has_vat:
            return idx
        if not include_vat and excludes_vat:
            return idx
    return None


def is_project_number(value: Any) -> bool:
    text = clean_text(value)
    if not text:
        return False
    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return False
    return number.is_integer() and 1 <= int(number) <= 999


def checked(value: Any) -> bool:
    text = clean_text(value)
    return text in {"✓", "✔", "√", "1", "true", "TRUE"} or text.lower() in {"yes", "y"}


def infer_status_label(headers: list[Any], row: list[Any], fallback: str) -> str:
    for idx, header in enumerate(headers):
        if idx >= len(row) or not checked(row[idx]):
            continue
        label = clean_text(header)
        compact = norm_key(label)
        if "تفاوض" in compact or "قيد" in compact:
            return "تم التقديم وقيد التفاوض والترسية"
        if "لم" in compact and "توقيع" in compact:
            return "تم الترسية ولم يتم توقيع العقد"
        if "توقيع" in compact:
            return "تم الترسية وتم توقيع العقد"
    return fallback


def workbook_sheets(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    sheets = []
    for ws in workbook.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            values = [clean_text(cell) for cell in row]
            if any(values):
                rows.append(values)
        sheets.append({"name": ws.title, "rows": rows})
    return sheets


def detect_header_row(rows: list[list[Any]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:25]):
        filled = sum(1 for cell in row if clean_text(cell))
        keywords = sum(
            1
            for cell in row
            if any(k in clean_text(cell) for k in ["المشروع", "المناقصة", "العميل", "المالك", "القيمة", "الحالة", "المحفظة"])
        )
        score = filled + (keywords * 3)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def sheet_objects(rows: list[list[Any]]) -> list[dict[str, str]]:
    if not rows:
        return []
    header_idx = detect_header_row(rows)
    headers = [clean_text(h) or f"column_{i + 1}" for i, h in enumerate(rows[header_idx])]
    objects = []
    for row in rows[header_idx + 1 :]:
        item = {}
        for idx, header in enumerate(headers):
            item[header] = clean_text(row[idx]) if idx < len(row) else ""
        if any(item.values()):
            objects.append(item)
    return objects


def status_code(label: str) -> str:
    text = label.strip()
    if not text:
        return "unknown"
    if "لم" in text and ("توقيع" in text or "توقيع العقد" in text):
        return "awarded_not_signed"
    if "توقيع" in text or "العقد" in text:
        return "awarded_signed"
    if "تفاوض" in text or "ترسية" in text or "قيد" in text:
        return "submitted_negotiation"
    return "unknown"


def build_portfolio_json(workbook_bytes: bytes, source_url: str) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheets = workbook_sheets(workbook)
    chosen = max(sheets, key=lambda sheet: len(sheet["rows"]), default={"name": "", "rows": []})
    rows = chosen["rows"]
    if not rows:
        raise RuntimeError("Portfolio workbook does not contain readable rows.")

    header_idx = detect_header_row(rows)
    headers = rows[header_idx]
    subheaders = rows[header_idx + 1] if header_idx + 1 < len(rows) else []
    has_status_subheaders = any(
        any(word in clean_text(cell) for word in ("توقيع", "تفاوض", "ترسية", "العقد"))
        for cell in subheaders
    )
    status_headers = subheaders if has_status_subheaders else headers
    data_rows = rows[header_idx + (2 if has_status_subheaders else 1) :]

    number_col = find_col(headers, "رقم", "م", fallback=0)
    project_col = find_col(headers, "المشروع", "المنافسة", fallback=2)
    client_col = find_col(headers, "العميل", "المالك", fallback=3)
    status_col = find_col(headers, "الحالة", "الترسية", fallback=4)
    portfolio_col = find_col(headers, "المحفظة", "القطاع", fallback=5)
    amount_ex_col = find_amount_col(headers, include_vat=False)
    amount_in_col = find_amount_col(headers, include_vat=True)
    if amount_ex_col is None:
        amount_ex_col = find_col(headers, "القيمة", "المبلغ", fallback=1)

    projects = []
    for index, row in enumerate(data_rows, start=1):
        if not any(clean_text(cell) for cell in row):
            continue
        if number_col is not None and number_col < len(row) and not is_project_number(row[number_col]):
            continue
        project_name = clean_text(row[project_col]) if project_col is not None and project_col < len(row) else ""
        amount_excl = number_value(row[amount_ex_col]) if amount_ex_col is not None and amount_ex_col < len(row) else 0.0
        if not project_name and not amount_excl:
            continue
        amount_incl = (
            number_value(row[amount_in_col])
            if amount_in_col is not None and amount_in_col < len(row)
            else amount_excl * (1 + VAT_RATE)
        )
        status_label = clean_text(row[status_col]) if status_col is not None and status_col < len(row) else ""
        status_label = infer_status_label(status_headers, row, status_label)
        projects.append(
            {
                "number": clean_text(row[number_col]) if number_col is not None and number_col < len(row) else str(index),
                "amount": amount_excl,
                "amountExclVat": amount_excl,
                "amountInclVat": amount_incl,
                "project": project_name or "-",
                "client": clean_text(row[client_col]) if client_col is not None and client_col < len(row) else "-",
                "status": status_code(status_label),
                "statusLabel": status_label or "غير محدد",
                "portfolio": clean_text(row[portfolio_col]) if portfolio_col is not None and portfolio_col < len(row) else "غير محدد",
            }
        )

    total_excl = sum(item["amountExclVat"] for item in projects)
    total_incl = sum(item["amountInclVat"] for item in projects)
    return {
        "source": "SharePoint Excel",
        "sourceUrl": source_url,
        "sheet": chosen["name"],
        "generatedAt": now_iso(),
        "currency": "SAR",
        "vatRate": VAT_RATE,
        "summary": {
            "projectCount": len(projects),
            "totalAmountExclVat": total_excl,
            "totalAmountInclVat": total_incl,
        },
        "projects": projects,
    }


def build_executive_json(workbook_bytes: bytes, source_url: str) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheets = workbook_sheets(workbook)
    return {
        "source": "SharePoint Excel",
        "sourceUrl": source_url,
        "generatedAt": now_iso(),
        "sheets": [
            {
                "name": sheet["name"],
                "rows": sheet["rows"],
                "objects": sheet_objects(sheet["rows"]),
            }
            for sheet in sheets
        ],
    }


def write_json(path: str, data: dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    log.info("Wrote %s", path)


def main() -> None:
    token = graph_token()
    portfolio_url = os.getenv("SP_PORTFOLIO_EXCEL_URL") or DEFAULT_PORTFOLIO_URL
    executive_url = os.getenv("SP_EXECUTIVE_EXCEL_URL") or DEFAULT_EXECUTIVE_URL

    portfolio_bytes = download_shared_workbook(portfolio_url, token)
    executive_bytes = download_shared_workbook(executive_url, token)

    write_json("portfolio/data.json", build_portfolio_json(portfolio_bytes, portfolio_url))
    write_json("executive-report/data.json", build_executive_json(executive_bytes, executive_url))


if __name__ == "__main__":
    main()
